from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import login, authenticate
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, Count
from django.http import JsonResponse
from django.utils import timezone
from datetime import date, datetime, timedelta
from .models import Employee, Department, Position, News, BonusTransfer, Holiday, StaffMember, Notification, SystemSettings
from .forms import EmployeeRegistrationForm, EmployeeLoginForm, BonusTransferForm, ProfileEditForm, NewsForm, StaffMemberForm
from django.utils import timezone
from django.http import HttpResponse
from django.template.loader import render_to_string
import openpyxl
from openpyxl.styles import Font, Alignment
from calendar import monthrange
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import mm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import os


def login_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = EmployeeLoginForm(request.POST)
        if form.is_valid():
            login_value = form.cleaned_data['login']
            password = form.cleaned_data['password']
            
            # Попытка входа по email или телефону
            try:
                if '@' in login_value:
                    user = Employee.objects.get(email=login_value)
                else:
                    user = Employee.objects.get(phone=login_value)
                
                user = authenticate(request, username=user.username, password=password)
                if user is not None:
                    login(request, user)
                    user.reset_monthly_balance()  # Проверка и сброс баланса
                    return redirect('home')
                else:
                    messages.error(request, 'Неверный пароль')
            except Employee.DoesNotExist:
                messages.error(request, 'Пользователь с таким email или телефоном не найден')
    else:
        form = EmployeeLoginForm()
    
    return render(request, 'employees/login.html', {'form': form})


def register_view(request):
    if request.user.is_authenticated:
        return redirect('home')
    
    if request.method == 'POST':
        form = EmployeeRegistrationForm(request.POST, request.FILES)
        if form.is_valid():
            user = form.save()
            login(request, user)
            
            # Проверяем, были ли начислены бонусы до регистрации
            if user.received_bonus_balance > 0:
                messages.success(
                    request, 
                    f'Регистрация успешна! Вам начислено {user.received_bonus_balance} руб. бонусов, которые были переведены до вашей регистрации.'
                )
            else:
                messages.success(request, 'Регистрация успешна!')
            return redirect('home')
    else:
        form = EmployeeRegistrationForm()
    
    return render(request, 'employees/register.html', {'form': form})


@login_required
def home_view(request):
    # Проверка и сброс баланса
    request.user.reset_monthly_balance()
    
    # Новости (последние 5)
    news = News.objects.all()[:5]
    
    # Рейтинг за текущий месяц
    current_month = timezone.now().month
    current_year = timezone.now().year
    rating = Employee.objects.filter(participates_in_bonus=True).annotate(
        total_received=Sum(
            'received_transfers__amount',
            filter=Q(received_transfers__created_at__month=current_month,
                    received_transfers__created_at__year=current_year,
                    received_transfers__is_deleted=False)
        )
    ).filter(total_received__isnull=False).order_by('-total_received')[:10]
    
    # Новые отзывы (последние 5) - исключаем удаленные
    recent_reviews = BonusTransfer.objects.filter(is_deleted=False).select_related('from_employee', 'to_employee').order_by('-created_at')[:5]
    
    # Непрочитанные переводы (уведомления)
    unread_transfers = BonusTransfer.objects.filter(
        to_employee=request.user,
        notification_sent=False
    ).select_related('from_employee').order_by('-created_at')
    
    # Помечаем переводы как прочитанные
    if unread_transfers.exists():
        BonusTransfer.objects.filter(
            to_employee=request.user,
            notification_sent=False
        ).update(notification_sent=True)
    
    # Проверка дней рождения и праздников
    today = date.today()
    birthdays = Employee.objects.filter(
        birth_date__month=today.month,
        birth_date__day=today.day
    ).exclude(id=request.user.id)
    
    holidays = Holiday.objects.filter(
        date=today,
        is_annual=True
    )
    
    # Дни рождения в этом месяце (для новостей)
    month_birthdays = StaffMember.objects.filter(
        birth_date__month=current_month,
        is_active=True
    ).select_related('department', 'position').order_by('birth_date__day')
    
    context = {
        'news': news,
        'rating': rating,
        'recent_reviews': recent_reviews,
        'unread_transfers': unread_transfers,
        'birthdays': birthdays,
        'holidays': holidays,
        'month_birthdays': month_birthdays,
        'current_month': current_month,
    }
    
    return render(request, 'employees/home.html', context)


@login_required
def employees_list_view(request):
    # Используем модель StaffMember вместо Employee - это сотрудники компании, не зарегистрированные в системе
    employees = StaffMember.objects.select_related('department', 'position', 'employee_profile').filter(is_active=True)
    
    # Фильтрация
    department_filter = request.GET.get('department')
    position_filter = request.GET.get('position')
    search_query = request.GET.get('search')
    
    if department_filter:
        employees = employees.filter(department_id=department_filter)
    if position_filter:
        employees = employees.filter(position_id=position_filter)
    if search_query:
        employees = employees.filter(
            Q(first_name__icontains=search_query) |
            Q(last_name__icontains=search_query) |
            Q(middle_name__icontains=search_query) |
            Q(email__icontains=search_query) |
            Q(phone__icontains=search_query)
        )
    
    departments = Department.objects.all()
    # Фильтруем должности только с указанным отделом, если выбран отдел
    if department_filter:
        positions = Position.objects.filter(department_id=department_filter, department__isnull=False)
    else:
        positions = Position.objects.filter(department__isnull=False)
    
    context = {
        'employees': employees,
        'departments': departments,
        'positions': positions,
        'selected_department': department_filter,
        'selected_position': position_filter,
        'search_query': search_query,
    }
    
    return render(request, 'employees/employees_list.html', context)


@login_required
def reviews_list_view(request):
    # Исключаем удаленные переводы
    reviews = BonusTransfer.objects.filter(is_deleted=False).select_related('from_employee', 'to_employee').all()
    
    # Фильтрация
    employee_filter = request.GET.get('employee')
    month_filter = request.GET.get('month')
    my_reviews = request.GET.get('my_reviews')
    
    if my_reviews == 'sent':
        reviews = reviews.filter(from_employee=request.user)
    elif my_reviews == 'received':
        reviews = reviews.filter(to_employee=request.user)
    
    if employee_filter:
        reviews = reviews.filter(
            Q(from_employee_id=employee_filter) | Q(to_employee_id=employee_filter)
        )
    
    if month_filter:
        year, month = month_filter.split('-')
        month = int(month)
        year = int(year)
        reviews = reviews.filter(created_at__month=month, created_at__year=year)
    
    employees = Employee.objects.all()
    
    # Генерация списка месяцев для фильтра
    months = []
    for i in range(12):
        date_obj = timezone.now() - timedelta(days=30*i)
        months.append(date_obj.strftime('%Y-%m'))
    
    context = {
        'reviews': reviews,
        'employees': employees,
        'months': months,
        'selected_employee': employee_filter,
        'selected_month': month_filter,
        'my_reviews': my_reviews,
    }
    
    return render(request, 'employees/reviews_list.html', context)


@login_required
def rating_view(request):
    period_type = request.GET.get('period', 'month')  # month, quarter, year
    period_value = request.GET.get('period_value')
    
    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    
    now = timezone.now()
    current_year = now.year
    current_month = now.month
    current_quarter = (current_month - 1) // 3 + 1
    
    # Определяем период для фильтрации
    if period_type == 'year':
        if period_value:
            year = int(period_value)
        else:
            year = current_year
        start_date = date(year, 1, 1)
        end_date = date(year, 12, 31)
        period_label = f'{year} год'
    elif period_type == 'quarter':
        if period_value:
            year, quarter = map(int, period_value.split('-Q'))
        else:
            year = current_year
            quarter = current_quarter
        start_month = (quarter - 1) * 3 + 1
        end_month = quarter * 3
        start_date = date(year, start_month, 1)
        last_day = monthrange(year, end_month)[1]
        end_date = date(year, end_month, last_day)
        period_label = f'{quarter} квартал {year}'
    else:  # month
        if period_value:
            year, month = map(int, period_value.split('-'))
        else:
            year = current_year
            month = current_month
        start_date = date(year, month, 1)
        last_day = monthrange(year, month)[1]
        end_date = date(year, month, last_day)
        period_label = f'{month_names_ru[month]} {year}'
    
    # Фильтруем переводы по периоду
    rating = Employee.objects.filter(participates_in_bonus=True).annotate(
        total_received=Sum(
            'received_transfers__amount',
            filter=Q(received_transfers__created_at__date__gte=start_date,
                    received_transfers__created_at__date__lte=end_date,
                    received_transfers__is_deleted=False)
        )
    ).filter(total_received__isnull=False).order_by('-total_received')
    
    # Данные для диаграммы
    chart_data = {
        'labels': [emp.get_full_name() for emp in rating[:10]],
        'data': [float(emp.total_received or 0) for emp in rating[:10]],
    }
    
    # Генерация списков для фильтров
    months = []
    for i in range(12):
        date_obj = now - timedelta(days=30*i)
        month_num = date_obj.month
        months.append({
            'value': date_obj.strftime('%Y-%m'),
            'label': f'{month_names_ru[month_num]} {date_obj.year}'
        })
    
    quarters = []
    for year in range(current_year - 2, current_year + 1):
        for q in range(1, 5):
            quarters.append({
                'value': f'{year}-Q{q}',
                'label': f'{q} квартал {year}'
            })
    
    years = []
    for year in range(current_year - 2, current_year + 1):
        years.append({
            'value': str(year),
            'label': f'{year} год'
        })
    
    context = {
        'rating': rating,
        'chart_data': chart_data,
        'months': months,
        'quarters': quarters,
        'years': years,
        'period_type': period_type,
        'period_value': period_value or (f'{current_year}-{current_month:02d}' if period_type == 'month' else f'{current_year}-Q{current_quarter}' if period_type == 'quarter' else str(current_year)),
        'period_label': period_label,
    }
    
    return render(request, 'employees/rating.html', context)


@login_required
def bonus_transfer_view(request):
    request.user.reset_monthly_balance()
    
    # Проверка участия в системе премирования
    if not request.user.participates_in_bonus:
        messages.error(request, 'Вы не участвуете в системе премирования')
        return redirect('home')
    
    # Получаем предвыбранного сотрудника из GET параметра
    preselected_staff_id = request.GET.get('staff_id')
    preselected_employee = None
    if preselected_staff_id:
        try:
            staff = StaffMember.objects.get(id=preselected_staff_id)
            if staff.employee_profile:
                preselected_employee = staff.employee_profile
            else:
                # Ищем Employee по совпадению ФИО
                full_name_parts = [staff.last_name, staff.first_name, staff.middle_name]
                full_name = ' '.join([p for p in full_name_parts if p]).strip()
                try:
                    employee = Employee.objects.get(
                        last_name=staff.last_name,
                        first_name=staff.first_name,
                        middle_name=staff.middle_name or ''
                    )
                    preselected_employee = employee
                except (Employee.DoesNotExist, Employee.MultipleObjectsReturned):
                    # Если не найдено или несколько, пробуем найти по полному имени
                    employees = Employee.objects.filter(
                        last_name=staff.last_name,
                        first_name=staff.first_name
                    )
                    if employees.exists():
                        preselected_employee = employees.first()
        except StaffMember.DoesNotExist:
            pass
    
    if request.method == 'POST':
        # Получаем staff_id из формы, если передан
        staff_id = request.POST.get('staff_id')
        to_employee_id = request.POST.get('to_employee')
        
        # Если передан staff_id, находим или создаем Employee по ФИО из StaffMember
        to_employee = None
        if staff_id:
            try:
                staff = StaffMember.objects.get(id=staff_id)
                # Проверяем участие сотрудника в системе премирования
                if not staff.participates_in_bonus:
                    messages.error(request, 'Этот сотрудник не участвует в системе премирования')
                    return redirect('bonus_transfer')
                
                # Ищем Employee по ФИО
                if staff.employee_profile:
                    to_employee = staff.employee_profile
                else:
                    # Ищем по ФИО
                    try:
                        to_employee = Employee.objects.get(
                            last_name=staff.last_name,
                            first_name=staff.first_name,
                            middle_name=staff.middle_name or ''
                        )
                    except Employee.DoesNotExist:
                        # Создаем нового Employee по ФИО из StaffMember автоматически
                        base_username = (staff.email or f"{staff.last_name}_{staff.first_name}").lower().replace(' ', '_').replace('-', '_')
                        username = base_username
                        counter = 1
                        while Employee.objects.filter(username=username).exists():
                            username = f"{base_username}_{counter}"
                            counter += 1
                        
                        # Генерируем уникальный телефон, если его нет
                        phone = staff.phone
                        if not phone:
                            # Генерируем уникальный временный телефон на основе ID и данных
                            import hashlib
                            phone_hash = hashlib.md5(f"{staff.id}_{staff.last_name}_{staff.first_name}".encode()).hexdigest()[:9]
                            phone = f'+7999{phone_hash}'
                            # Проверяем уникальность
                            counter = 1
                            while Employee.objects.filter(phone=phone).exists():
                                phone = f'+7999{phone_hash[:8]}{counter}'
                                counter += 1
                        else:
                            # Проверяем, не используется ли телефон другим активным пользователем
                            if Employee.objects.filter(phone=phone, is_active=True).exists():
                                # Если телефон занят активным пользователем, генерируем новый
                                import hashlib
                                phone_hash = hashlib.md5(f"{staff.id}_{staff.last_name}_{staff.first_name}".encode()).hexdigest()[:9]
                                phone = f'+7999{phone_hash}'
                                counter = 1
                                while Employee.objects.filter(phone=phone).exists():
                                    phone = f'+7999{phone_hash[:8]}{counter}'
                                    counter += 1
                        
                        # Генерируем уникальный email, если его нет
                        email = staff.email
                        if not email:
                            email = f"{staff.last_name.lower()}.{staff.first_name.lower()}@company.local"
                            counter = 1
                            while Employee.objects.filter(email=email).exists():
                                email = f"{staff.last_name.lower()}.{staff.first_name.lower()}{counter}@company.local"
                                counter += 1
                        else:
                            # Проверяем, не используется ли email другим активным пользователем
                            if Employee.objects.filter(email=email, is_active=True).exists():
                                # Если email занят активным пользователем, генерируем новый
                                email = f"{staff.last_name.lower()}.{staff.first_name.lower()}@company.local"
                                counter = 1
                                while Employee.objects.filter(email=email).exists():
                                    email = f"{staff.last_name.lower()}.{staff.first_name.lower()}{counter}@company.local"
                                    counter += 1
                        
                        to_employee = Employee.objects.create(
                            username=username,
                            last_name=staff.last_name,
                            first_name=staff.first_name,
                            middle_name=staff.middle_name or '',
                            email=email,
                            phone=phone,
                            department=staff.department,
                            position=staff.position,
                            photo=staff.photo,
                            participates_in_bonus=staff.participates_in_bonus,
                            is_active=False,  # Неактивный, так как не зарегистрирован
                            monthly_bonus_balance=staff.monthly_bonus_amount or 1000,  # Не получает месячные бонусы автоматически
                            received_bonus_balance=0  # Начальный баланс полученных бонусов
                        )
                        # Связываем StaffMember с Employee
                        staff.employee_profile = to_employee
                        staff.save()
                    except Employee.MultipleObjectsReturned:
                        to_employee = Employee.objects.filter(
                            last_name=staff.last_name,
                            first_name=staff.first_name,
                            middle_name=staff.middle_name or ''
                        ).first()
            except StaffMember.DoesNotExist:
                messages.error(request, 'Сотрудник не найден в справочнике')
                return redirect('bonus_transfer')
        
        # Если to_employee не найден через staff_id, используем стандартную форму
        if not to_employee and to_employee_id:
            try:
                to_employee = Employee.objects.get(id=to_employee_id)
            except Employee.DoesNotExist:
                messages.error(request, 'Сотрудник не найден')
                return redirect('bonus_transfer')
        
        # Проверяем, что выбран сотрудник
        if not to_employee:
            messages.error(request, 'Не выбран сотрудник. Выберите сотрудника из списка выше.')
            return redirect('bonus_transfer')
        
        # Создаем форму с найденным сотрудником
        form_data = request.POST.copy()
        form_data['to_employee'] = to_employee.id
        form = BonusTransferForm(form_data, request.FILES, from_employee=request.user)
        
        if form.is_valid():
            transfer = form.save(commit=False)
            transfer.from_employee = request.user
            transfer.to_employee = to_employee
            
            # Проверка баланса
            if transfer.amount > request.user.monthly_bonus_balance:
                messages.error(request, 'Недостаточно средств')
                return redirect('bonus_transfer')
            
            # Проверка участия в системе премирования
            if not to_employee.participates_in_bonus:
                messages.error(request, 'Этот сотрудник не участвует в системе премирования')
                return redirect('bonus_transfer')
            
            if not request.user.participates_in_bonus:
                messages.error(request, 'Вы не участвуете в системе премирования')
                return redirect('bonus_transfer')
            
            # Перевод
            request.user.monthly_bonus_balance -= transfer.amount
            request.user.save()
            
            to_employee.received_bonus_balance += transfer.amount
            to_employee.save()
            
            transfer.save()
            
            # Создаем уведомление для получателя (если он зарегистрирован)
            if to_employee.is_active:
                Notification.objects.create(
                    user=to_employee,
                    type='transfer_received',
                    title='Получен перевод бонусов',
                    message=f'Вы получили {transfer.amount} руб. от {request.user.get_full_name()}. Причина: {transfer.get_reason_display()}',
                    related_transfer=transfer
                )
            
            messages.success(request, f'Премия успешно переведена {to_employee.get_full_name()}!')
            return redirect('bonus_transfer')
    else:
        form = BonusTransferForm(from_employee=request.user, preselected_employee=preselected_employee)
    
    # Получаем ВСЕХ сотрудников из справочника (все, кто участвует в системе премирования)
    all_staff = StaffMember.objects.filter(
        is_active=True,
        participates_in_bonus=True
    ).select_related('department', 'position', 'employee_profile').order_by('department__name', 'position__name', 'last_name', 'first_name')
    
    # Показываем всех сотрудников из справочника, даже если у них нет Employee
    staff_list = []
    for staff in all_staff:
        # Пропускаем только текущего пользователя, если он есть в списке
        if staff.employee_profile and staff.employee_profile.id == request.user.id:
            continue
        
        # Находим Employee по ФИО (если есть), но показываем всех из справочника
        employee = None
        if staff.employee_profile:
            employee = staff.employee_profile
        else:
            # Ищем Employee по совпадению ФИО
            try:
                employee = Employee.objects.get(
                    last_name=staff.last_name,
                    first_name=staff.first_name,
                    middle_name=staff.middle_name or ''
                )
            except (Employee.DoesNotExist, Employee.MultipleObjectsReturned):
                # Если не найдено точно, пробуем найти по фамилии и имени
                employees = Employee.objects.filter(
                    last_name=staff.last_name,
                    first_name=staff.first_name
                ).exclude(id=request.user.id)
                if employees.exists():
                    employee = employees.first()
        
        # Добавляем всех сотрудников из справочника
        staff_list.append({
            'staff': staff,
            'employee': employee  # Может быть None, если Employee не найден
        })
    
    departments = Department.objects.all().order_by('name')
    positions = Position.objects.all().order_by('name')
    
    context = {
        'form': form,
        'balance': request.user.monthly_bonus_balance,
        'all_staff': staff_list,
        'departments': departments,
        'positions': positions,
        'preselected_staff_id': preselected_staff_id,
    }
    
    return render(request, 'employees/bonus_transfer.html', context)


@login_required
def profile_view(request):
    if request.method == 'POST':
        form = ProfileEditForm(request.POST, request.FILES, instance=request.user)
        if form.is_valid():
            form.save()
            # Обновляем фото в справочнике по ФИО (если ФИО совпадает)
            if request.user.photo:
                # Сначала ищем по прямой связи
                staff_member = StaffMember.objects.filter(employee_profile=request.user).first()
                if not staff_member:
                    # Если прямой связи нет, ищем по ФИО
                    staff_member = StaffMember.objects.filter(
                        last_name=request.user.last_name,
                        first_name=request.user.first_name,
                        middle_name=request.user.middle_name or ''
                    ).first()
                
                if staff_member:
                    staff_member.photo = request.user.photo
                    staff_member.save()
                    messages.success(request, 'Профиль и фото в справочнике успешно обновлены!')
                else:
                    messages.success(request, 'Профиль успешно обновлен!')
            else:
                messages.success(request, 'Профиль успешно обновлен!')
            return redirect('profile')
    else:
        form = ProfileEditForm(instance=request.user)
    
    # Проверяем, есть ли сотрудник в справочнике (по ФИО)
    staff_member_in_directory = None
    staff_member = StaffMember.objects.filter(employee_profile=request.user).first()
    if not staff_member:
        staff_member = StaffMember.objects.filter(
            last_name=request.user.last_name,
            first_name=request.user.first_name,
            middle_name=request.user.middle_name or ''
        ).first()
    if staff_member:
        staff_member_in_directory = staff_member
    
    # Отзывы пользователя
    sent_reviews = BonusTransfer.objects.filter(from_employee=request.user).order_by('-created_at')
    received_reviews = BonusTransfer.objects.filter(to_employee=request.user).order_by('-created_at')
    
    context = {
        'form': form,
        'sent_reviews': sent_reviews,
        'received_reviews': received_reviews,
        'staff_member_in_directory': staff_member_in_directory,
    }
    
    return render(request, 'employees/profile.html', context)


@login_required
def logout_view(request):
    from django.contrib.auth import logout
    logout(request)
    return redirect('login')


def get_positions_by_department(request):
    """AJAX-эндпоинт для получения должностей по отделу"""
    department_id = request.GET.get('department_id')
    if department_id:
        # Фильтруем только должности с указанным отделом (исключаем null)
        positions = Position.objects.filter(department_id=department_id, department__isnull=False).values('id', 'name')
        return JsonResponse(list(positions), safe=False)
    return JsonResponse([], safe=False)


@login_required
def notifications_view(request):
    """Страница уведомлений"""
    notifications = Notification.objects.filter(user=request.user).order_by('-created_at')
    unread_count = notifications.filter(is_read=False).count()
    
    # Помечаем все как прочитанные при явном запросе
    if request.method == 'GET' and 'mark_read' in request.GET:
        Notification.objects.filter(user=request.user, is_read=False).update(is_read=True)
        return redirect('notifications')
    
    context = {
        'notifications': notifications,
        'unread_count': unread_count,
    }
    
    return render(request, 'employees/notifications.html', context)


@login_required
def admin_panel_view(request):
    """Панель администратора"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа к панели администратора')
        return redirect('home')
    
    context = {
        'staff_count': StaffMember.objects.filter(is_active=True).count(),
        'transfers_count': BonusTransfer.objects.filter(is_deleted=False).count(),
        'news_count': News.objects.count(),
    }
    
    return render(request, 'employees/admin_panel.html', context)


@login_required
def admin_staff_manage_view(request):
    """Управление сотрудниками для администратора"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    if request.method == 'POST':
        form = StaffMemberForm(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            messages.success(request, 'Сотрудник успешно добавлен!')
            return redirect('admin_staff_manage')
    else:
        form = StaffMemberForm()
    
    # Все сотрудники — фильтрация на клиенте (как при начислении премии)
    staff_members = StaffMember.objects.select_related('department', 'position', 'employee_profile').order_by('last_name', 'first_name')
    
    context = {
        'form': form,
        'staff_members': staff_members,
    }
    
    return render(request, 'employees/admin_staff_manage.html', context)


@login_required
def admin_staff_edit_view(request, staff_id):
    """Редактирование сотрудника администратором"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    staff_member = get_object_or_404(StaffMember, id=staff_id)
    
    if request.method == 'POST':
        form = StaffMemberForm(request.POST, request.FILES, instance=staff_member)
        if form.is_valid():
            form.save()
            messages.success(request, 'Данные сотрудника обновлены!')
            return redirect('admin_staff_manage')
    else:
        form = StaffMemberForm(instance=staff_member)
    
    context = {
        'form': form,
        'staff_member': staff_member,
    }
    
    return render(request, 'employees/admin_staff_edit.html', context)


@login_required
def admin_staff_delete_view(request, staff_id):
    """Удаление сотрудника из справочника администратором"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    staff_member = get_object_or_404(StaffMember, id=staff_id)
    staff_member.delete()
    messages.success(request, 'Сотрудник удален из справочника')
    return redirect('admin_staff_manage')


@login_required
def admin_user_delete_view(request, employee_id):
    """Удаление учётной записи пользователя без удаления из справочника сотрудников"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    if employee_id == request.user.id:
        messages.error(request, 'Нельзя удалить свою учётную запись')
        return redirect('admin_staff_manage')
    
    employee = get_object_or_404(Employee, id=employee_id)
    full_name = employee.get_full_name()
    
    # Связь StaffMember.employee_profile имеет on_delete=SET_NULL — при удалении Employee
    # запись в справочнике останется, связь обнулится автоматически
    employee.delete()
    messages.success(request, f'Учётная запись {full_name} удалена. Сотрудник остаётся в справочнике.')
    
    # Возврат на страницу, с которой пришли (управление сотрудниками или администраторами)
    next_url = request.GET.get('next')
    if next_url and 'manage-admins' in next_url:
        return redirect('admin_manage_admins')
    return redirect('admin_staff_manage')


@login_required
def admin_transfers_view(request):
    """Управление переводами для администратора"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    transfers = BonusTransfer.objects.filter(is_deleted=False).select_related('from_employee', 'to_employee').order_by('-created_at')
    
    context = {
        'transfers': transfers,
    }
    
    return render(request, 'employees/admin_transfers.html', context)


@login_required
def admin_transfer_delete_view(request, transfer_id):
    """Удаление перевода администратором"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    transfer = get_object_or_404(BonusTransfer, id=transfer_id)
    
    if request.method == 'POST':
        # Возвращаем средства
        transfer.from_employee.monthly_bonus_balance += transfer.amount
        transfer.from_employee.save()
        
        transfer.to_employee.received_bonus_balance -= transfer.amount
        if transfer.to_employee.received_bonus_balance < 0:
            transfer.to_employee.received_bonus_balance = 0
        transfer.to_employee.save()
        
        # Помечаем как удаленный
        transfer.is_deleted = True
        transfer.deleted_by = request.user
        transfer.deleted_at = timezone.now()
        transfer.save()
        
        # Создаем уведомление для получателя
        Notification.objects.create(
            user=transfer.to_employee,
            type='transfer_cancelled',
            title='Перевод отменен',
            message=f'Перевод на сумму {transfer.amount} руб. от {transfer.from_employee.get_full_name()} был отменен администратором.',
            related_transfer=transfer
        )
        
        # Создаем уведомление для отправителя
        Notification.objects.create(
            user=transfer.from_employee,
            type='transfer_cancelled',
            title='Перевод отменен',
            message=f'Ваш перевод на сумму {transfer.amount} руб. для {transfer.to_employee.get_full_name()} был отменен администратором. Средства возвращены на ваш баланс.',
            related_transfer=transfer
        )
        
        messages.success(request, 'Перевод удален, средства возвращены')
        return redirect('admin_transfers')
    
    context = {
        'transfer': transfer,
    }
    
    return render(request, 'employees/admin_transfer_delete.html', context)


@login_required
def admin_news_create_view(request):
    """Создание новостей администратором"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    if request.method == 'POST':
        form = NewsForm(request.POST)
        if form.is_valid():
            news = form.save(commit=False)
            news.author = request.user
            news.save()
            
            # Создаем уведомления для всех пользователей
            employees = Employee.objects.exclude(id=request.user.id)
            for employee in employees:
                Notification.objects.create(
                    user=employee,
                    type='news',
                    title='Новая новость',
                    message=f'{news.title}',
                )
            
            messages.success(request, 'Новость создана!')
            return redirect('home')
    else:
        form = NewsForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'employees/admin_news_create.html', context)


@login_required
def admin_bonus_participation_view(request):
    """Управление участием сотрудников в системе премирования"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        participates = request.POST.get('participates') == 'true'
        try:
            employee = Employee.objects.get(id=employee_id)
            employee.participates_in_bonus = participates
            employee.save()
            
            # Синхронизируем с StaffMember по ФИО
            staff_members = StaffMember.objects.filter(
                last_name=employee.last_name,
                first_name=employee.first_name,
                middle_name=employee.middle_name or ''
            )
            for staff in staff_members:
                staff.participates_in_bonus = participates
                staff.save()
            
            return JsonResponse({'success': True})
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Сотрудник не найден'})
    
    employees = Employee.objects.select_related('department', 'position').all().order_by('last_name', 'first_name')
    
    context = {
        'employees': employees,
    }
    
    return render(request, 'employees/admin_bonus_participation.html', context)


@login_required
def admin_settings_view(request):
    """Настройки системы для администратора"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    settings = SystemSettings.get_settings()
    
    if request.method == 'POST':
        form_type = request.POST.get('form_type')
        
        if form_type == 'bonus':
            monthly_bonus_amount = request.POST.get('monthly_bonus_amount')
            if monthly_bonus_amount:
                try:
                    settings.monthly_bonus_amount = float(monthly_bonus_amount)
                    settings.save()
                    messages.success(request, 'Сумма бонусных рублей обновлена!')
                except ValueError:
                    messages.error(request, 'Неверное значение суммы')
        
        elif form_type == 'logo':
            company_logo = request.FILES.get('company_logo')
            if company_logo:
                settings.company_logo = company_logo
                settings.save()
                messages.success(request, 'Логотип компании обновлен!')
            else:
                messages.warning(request, 'Файл не выбран')
        
        return redirect('admin_settings')
    
    context = {
        'settings': settings,
    }
    
    return render(request, 'employees/admin_settings.html', context)


@login_required
def admin_manage_admins_view(request):
    """Управление администраторами"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    if request.method == 'POST':
        employee_id = request.POST.get('employee_id')
        is_admin = request.POST.get('is_admin') == 'true'
        try:
            employee = Employee.objects.get(id=employee_id)
            employee.is_admin = is_admin
            employee.save()
            return JsonResponse({'success': True})
        except Employee.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Сотрудник не найден'})
    
    employees = Employee.objects.select_related('department', 'position').all().order_by('last_name', 'first_name')
    
    context = {
        'employees': employees,
    }
    
    return render(request, 'employees/admin_manage_admins.html', context)


@login_required
def admin_export_transfers_view(request):
    """Выгрузка реестра премий за месяц"""
    if not request.user.is_admin:
        messages.error(request, 'У вас нет прав доступа')
        return redirect('home')
    
    export_format = request.GET.get('format', 'excel')  # excel или pdf
    
    month_filter = request.GET.get('month')
    if month_filter:
        year, month = map(int, month_filter.split('-'))
    else:
        now = timezone.now()
        year = now.year
        month = now.month
    
    month_names_ru = {
        1: 'Январь', 2: 'Февраль', 3: 'Март', 4: 'Апрель',
        5: 'Май', 6: 'Июнь', 7: 'Июль', 8: 'Август',
        9: 'Сентябрь', 10: 'Октябрь', 11: 'Ноябрь', 12: 'Декабрь'
    }
    
    transfers = BonusTransfer.objects.filter(
        created_at__year=year,
        created_at__month=month,
        is_deleted=False
    ).select_related('to_employee', 'to_employee__department', 'to_employee__position').order_by('to_employee__last_name')
    if export_format == 'pdf':
        return _export_transfers_pdf(transfers, year, month, month_names_ru)
    else:
        return _export_transfers_excel(transfers, year, month, month_names_ru)


def _export_transfers_excel(transfers, year, month, month_names_ru):
    """Экспорт в Excel"""
    # Создаем Excel файл
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Реестр {month_names_ru[month]} {year}"
    
    # Шапка
    ws.merge_cells('A1:D1')
    ws['A1'] = f'Начисление из горизонтального премирования за "{month_names_ru[month].upper()}" {year}'
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal='center')
    
    # Заголовки таблицы
    headers = ['ФИО', 'Отдел', 'Должность', 'Сумма']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')
    
    # Данные
    row = 4
    total_amount = 0
    for transfer in transfers:
        employee = transfer.to_employee
        ws.cell(row=row, column=1, value=employee.get_full_name())
        ws.cell(row=row, column=2, value=employee.department.name if employee.department else '')
        ws.cell(row=row, column=3, value=employee.position.name if employee.position else '')
        ws.cell(row=row, column=4, value=float(transfer.amount))
        total_amount += float(transfer.amount)
        row += 1
    
    # Итого
    ws.cell(row=row, column=3, value='').font = Font(bold=True)
    ws.cell(row=row, column=4, value=total_amount).font = Font(bold=True)
    
    # Подпись генерального директора
    row += 2
    ws.cell(row=row, column=1, value='Генеральный директор')
    ws.cell(row=row, column=1).font = Font(bold=True)
    row += 3
    ws.cell(row=row, column=1, value='_________________')
    
    # Настройка ширины колонок
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 15
    
    # Создаем HTTP ответ
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="реестр_премий_{month_names_ru[month]}_{year}.xlsx"'
    wb.save(response)
    
    return response


def _export_transfers_pdf(transfers, year, month, month_names_ru):
    """Экспорт в PDF с поддержкой кириллицы"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="реестр_премий_{month_names_ru[month]}_{year}.pdf"'
    
    # Пробуем зарегистрировать шрифт с поддержкой кириллицы
    # Используем системные шрифты Windows или стандартные Unicode-шрифты
    font_name = 'Helvetica'  # По умолчанию
    
    # Пробуем найти и зарегистрировать шрифт с поддержкой кириллицы
    try:
        import platform
        system = platform.system()
        
        if system == 'Windows':
            # Пробуем использовать стандартные шрифты Windows
            font_paths = [
                'C:/Windows/Fonts/arial.ttf',
                'C:/Windows/Fonts/arialbd.ttf',
                'C:/Windows/Fonts/times.ttf',
            ]
        elif system == 'Linux':
            font_paths = [
                '/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf',
                '/usr/share/fonts/truetype/liberation/LiberationSans-Regular.ttf',
            ]
        else:
            font_paths = []
        
        for font_path in font_paths:
            if os.path.exists(font_path):
                try:
                    pdfmetrics.registerFont(TTFont('CyrillicFont', font_path, 'UTF-8'))
                    font_name = 'CyrillicFont'
                    break
                except:
                    continue
    except Exception as e:
        # Если не удалось зарегистрировать шрифт, используем стандартный
        pass
    
    doc = SimpleDocTemplate(response, pagesize=A4, rightMargin=20*mm, leftMargin=20*mm, topMargin=20*mm, bottomMargin=20*mm)
    story = []
    
    styles = getSampleStyleSheet()
    
    # Создаем стили с правильным шрифтом
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Normal'],
        fontSize=14,
        textColor=colors.HexColor('#000000'),
        spaceAfter=30,
        alignment=1,  # Center
        fontName=font_name,
        encoding='utf-8',
    )
    
    normal_style = ParagraphStyle(
        'CustomNormal',
        parent=styles['Normal'],
        fontName=font_name,
        fontSize=12,
        encoding='utf-8',
    )
    
    # Заголовок
    title_text = f'Начисление из горизонтального премирования за "{month_names_ru[month].upper()}" {year}'
    story.append(Paragraph(title_text, title_style))
    story.append(Spacer(1, 12))
    
    # Подготовка данных для таблицы
    data = [['ФИО', 'Отдел', 'Должность', 'Сумма']]
    total_amount = 0
    
    for transfer in transfers:
        employee = transfer.to_employee
        full_name = employee.get_full_name()
        dept_name = employee.department.name if employee.department else ''
        pos_name = employee.position.name if employee.position else ''
        
        data.append([
            full_name,
            dept_name,
            pos_name,
            f'{float(transfer.amount):.2f}'
        ])
        total_amount += float(transfer.amount)
    
    # Итого
    data.append(['', '', '', f'{total_amount:.2f}'])
    
    # Создание таблицы
    table = Table(data, colWidths=[60*mm, 60*mm, 60*mm, 25*mm])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#ffffff')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#000000')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),
        ('FONTNAME', (0, 0), (-1, 0), font_name),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -2), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTNAME', (0, -1), (-1, -1), font_name),
        ('FONTSIZE', (0, -1), (-1, -1), 12),
        ('FONTNAME', (0, 1), (-1, -2), font_name),
        ('FONTSIZE', (0, 1), (-1, -2), 10),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 20))

    # Подпись
    story.append(Paragraph('Генеральный директор', normal_style))
    story.append(Spacer(1, 30))
    story.append(Paragraph('_________________/_____________________', normal_style))
    
    doc.build(story)
    return response


@login_required
def update_staff_photo_view(request):
    """Обновление фото сотрудника в справочнике через профиль"""
    if request.method == 'POST':
        photo = request.FILES.get('photo')
        if photo:
            # Находим или создаем запись в справочнике для текущего пользователя
            staff_member, created = StaffMember.objects.get_or_create(
                employee_profile=request.user,
                defaults={
                    'first_name': request.user.first_name,
                    'last_name': request.user.last_name,
                    'middle_name': request.user.middle_name,
                    'email': request.user.email,
                    'phone': request.user.phone,
                    'department': request.user.department,
                    'position': request.user.position,
                }
            )
            staff_member.photo = photo
            staff_member.save()
            messages.success(request, 'Фото обновлено в справочнике!')
        return redirect('profile')
    
    return redirect('profile')